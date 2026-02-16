"""
EcoPOOL League - Excel Workbook Importer
Imports player roster, pairs, matchups, and scores from the master Excel scoresheet.
"""

import logging
from datetime import datetime
from typing import Optional

import openpyxl

from database import DatabaseManager

_logger = logging.getLogger(__name__)


class ExcelImporter:
    def __init__(self, db: DatabaseManager):
        self.db = db

    def import_workbook(self, filepath: str) -> tuple[bool, str]:
        """Import an entire Excel workbook into the database.

        Returns (success, message).
        """
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
        except Exception as e:
            return False, f"Failed to open workbook: {e}"

        sheet_names = wb.sheetnames
        if not sheet_names:
            return False, "Workbook has no sheets"

        # Find the roster sheet
        roster_ws = None
        for name in sheet_names:
            if 'roster' in name.lower() or 'stats' in name.lower():
                roster_ws = wb[name]
                break
        if roster_ws is None:
            roster_ws = wb[sheet_names[0]]

        # Import players from roster
        name_to_id = self._import_players_from_roster(roster_ws)
        player_count = len(name_to_id)

        # Ensure we have an active season
        season = self.db.get_active_season()
        if season is None:
            season_id = self.db.create_season("Season 26")
        else:
            season_id = season.id

        # Import week sheets
        weeks_imported = 0
        matches_imported = 0
        for name in sheet_names:
            if name.lower().startswith('week'):
                ws = wb[name]
                # Skip empty week sheets (check if there's partner data)
                if self._is_week_empty(ws):
                    continue
                count = self._import_week(ws, name, season_id, name_to_id)
                if count > 0:
                    weeks_imported += 1
                    matches_imported += count

        wb.close()

        return True, (
            f"Imported {player_count} players, "
            f"{weeks_imported} week(s), "
            f"{matches_imported} matches"
        )

    def _import_players_from_roster(self, ws) -> dict[str, int]:
        """Read the Roster & Stats sheet and create/update players.

        Returns a dict mapping 'First Last' -> player_id.
        """
        name_to_id = {}

        # Row 2 has headers, data starts at row 3
        for row in range(3, ws.max_row + 1):
            first = ws.cell(row=row, column=2).value  # B
            last = ws.cell(row=row, column=3).value    # C
            if not first or not last:
                continue

            first = str(first).strip()
            last = str(last).strip()
            full_name = f"{first} {last}"

            email = ws.cell(row=row, column=4).value or ""  # D
            venmo = ws.cell(row=row, column=5).value or ""  # E
            email = str(email).strip()
            venmo = str(venmo).strip()

            # Look up existing player by name
            player_id = self.db.find_player_by_name(full_name)
            if player_id is None:
                player_id = self.db.add_player(full_name, email=email, venmo=venmo)
                _logger.info(f"Created player: {full_name} (id={player_id})")
            else:
                # Update email/venmo if provided
                if email or venmo:
                    self.db.update_player(player_id, full_name, email=email, venmo=venmo)

            name_to_id[full_name] = player_id

        return name_to_id

    def _is_week_empty(self, ws) -> bool:
        """Check if a week sheet has no partner data (empty week)."""
        # Partners section starts around row 23. Check if there's any player name.
        for row in range(23, min(40, ws.max_row + 1)):
            val = ws.cell(row=row, column=2).value  # B column = first name
            if val and str(val).strip():
                return False
        return True

    def _import_week(self, ws, week_name: str, season_id: int,
                     name_to_id: dict[str, int]) -> int:
        """Import one week sheet (pairs, matchups, scores).

        Returns the number of matches imported.
        """
        # Extract a date from the week name if possible, otherwise use today
        date_str = datetime.now().strftime("%Y-%m-%d")

        # Check if a league night for this date already exists
        conn = self.db.get_connection()
        cursor = conn.cursor()
        cursor.execute(
            "SELECT id FROM league_nights WHERE notes = ? AND season_id = ?",
            (week_name, season_id)
        )
        existing = cursor.fetchone()
        if existing:
            _logger.info(f"League night for '{week_name}' already exists, skipping")
            return 0

        # Create league night
        ln_id = self.db.create_league_night(date_str, season_id=season_id)
        # Store week name in notes for idempotency
        cursor.execute("UPDATE league_nights SET notes = ? WHERE id = ?", (week_name, ln_id))
        conn.commit()

        # Parse pairs and their per-match scores
        pairs_data = self._parse_pairs(ws, ln_id, name_to_id)
        if not pairs_data:
            return 0

        # Parse matchup schedule
        matchups = self._parse_matchups(ws)
        if not matchups:
            return 0

        # Reconstruct games from matchups + pair scores
        match_count = self._reconstruct_games(pairs_data, matchups, ln_id, season_id)
        return match_count

    def _parse_pairs(self, ws, league_night_id: int,
                     name_to_id: dict[str, int]) -> dict[int, dict]:
        """Parse the Partners section of a week sheet.

        Returns a dict mapping team_number -> {
            'pair_id': int,
            'player1_id': int,
            'player2_id': int,
            'scores': [score1, score2, ...],  # per-match scores
            'match_counter': 0,  # tracks how many matches consumed
        }
        """
        pairs = {}

        # Partners section: rows 23-38, two rows per team
        # Odd rows: team#, first, last, match1-4 scores, total, wins, losses
        # Even rows: partner first, last
        row = 23
        while row <= min(38, ws.max_row):
            team_num = ws.cell(row=row, column=1).value  # A
            if team_num is None or not isinstance(team_num, (int, float)):
                row += 2
                continue

            team_num = int(team_num)

            # Player 1 (odd row)
            p1_first = ws.cell(row=row, column=2).value
            p1_last = ws.cell(row=row, column=3).value

            # Per-match scores (columns D-G)
            scores = []
            for col in range(4, 8):  # D=4, E=5, F=6, G=7
                val = ws.cell(row=row, column=col).value
                scores.append(int(val) if val is not None else 0)

            # Player 2 (even row)
            p2_first = ws.cell(row=row + 1, column=2).value
            p2_last = ws.cell(row=row + 1, column=3).value

            if not p1_first or not p1_last:
                row += 2
                continue

            p1_name = f"{str(p1_first).strip()} {str(p1_last).strip()}"
            p1_id = name_to_id.get(p1_name)
            if p1_id is None:
                _logger.warning(f"Player not found: {p1_name}")
                row += 2
                continue

            p2_id = None
            if p2_first and p2_last:
                p2_name = f"{str(p2_first).strip()} {str(p2_last).strip()}"
                p2_id = name_to_id.get(p2_name)
                if p2_id is None:
                    _logger.warning(f"Player not found: {p2_name}")

            # Create pair in DB
            pair_name = p1_name if p2_id is None else f"{p1_name} & {p2_name}"
            pair_id = self.db.create_pair(league_night_id, p1_id, p2_id, pair_name)

            pairs[team_num] = {
                'pair_id': pair_id,
                'player1_id': p1_id,
                'player2_id': p2_id,
                'scores': scores,
                'match_counter': 0,
            }

            row += 2

        return pairs

    def _parse_matchups(self, ws) -> list[tuple[int, int, int]]:
        """Parse the Matchups section of a week sheet.

        Returns a list of (set_number, team1_num, team2_num) in order.
        """
        matchups = []
        current_set = 0

        # Matchups section starts around row 42 (row 41 is header)
        for row in range(42, min(60, ws.max_row + 1)):
            # Set number in column A (may be merged/None for subsequent rows in same set)
            set_val = ws.cell(row=row, column=1).value
            if set_val is not None:
                try:
                    current_set = int(set_val)
                except (ValueError, TypeError):
                    continue

            team1 = ws.cell(row=row, column=2).value  # B
            team2 = ws.cell(row=row, column=3).value  # C

            if team1 is None or team2 is None:
                continue

            try:
                team1 = int(team1)
                team2 = int(team2)
            except (ValueError, TypeError):
                continue

            matchups.append((current_set, team1, team2))

        return matchups

    def _reconstruct_games(self, pairs_data: dict, matchups: list,
                           league_night_id: int, season_id: int) -> int:
        """Cross-reference matchup schedule with pair scores to create matches/games.

        For each matchup, uses each pair's match_counter as index into their scores list
        to get both teams' scores for that game.

        Returns number of matches created.
        """
        match_count = 0

        for set_num, team1_num, team2_num in matchups:
            pair1 = pairs_data.get(team1_num)
            pair2 = pairs_data.get(team2_num)

            if pair1 is None or pair2 is None:
                _logger.warning(f"Skipping matchup: team {team1_num} vs {team2_num} (pair not found)")
                continue

            # Get scores using each pair's match counter
            idx1 = pair1['match_counter']
            idx2 = pair2['match_counter']

            team1_score = pair1['scores'][idx1] if idx1 < len(pair1['scores']) else 0
            team2_score = pair2['scores'][idx2] if idx2 < len(pair2['scores']) else 0

            # Advance counters
            pair1['match_counter'] += 1
            pair2['match_counter'] += 1

            # Determine winner (higher score wins)
            if team1_score > team2_score:
                winner_team = 1
            elif team2_score > team1_score:
                winner_team = 2
            else:
                winner_team = 1  # Tie-break: team1 wins

            # Create match (best_of=1, already completed)
            match_id = self.db.create_match(
                team1_p1=pair1['player1_id'],
                team1_p2=pair1['player2_id'],
                team2_p1=pair2['player1_id'],
                team2_p2=pair2['player2_id'],
                best_of=1,
                league_night_id=league_night_id,
                pair1_id=pair1['pair_id'],
                pair2_id=pair2['pair_id'],
                status='completed',
                season_id=season_id,
                round_number=set_num,
            )

            # Mark match complete
            conn = self.db.get_connection()
            cursor = conn.cursor()
            cursor.execute(
                "UPDATE matches SET is_complete = 1 WHERE id = ?", (match_id,)
            )
            conn.commit()

            # Create game with scores
            game_id = self.db.create_game(match_id, game_number=1)
            self.db.update_game(
                game_id,
                team1_score=team1_score,
                team2_score=team2_score,
                team1_group='',
                balls_pocketed={},
                winner_team=winner_team,
            )

            match_count += 1

        return match_count
